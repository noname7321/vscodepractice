import openpyxl
# 创建一个新的工作簿
workbook = openpyxl.Workbook() 
# 选择默认的工作表
sheet = workbook.active
sheet = workbook.create_sheet(title='MySheet')  # 创建一个新的工作表并命名为"MySheet"
# 在A1单元格中写入数据
sheet['A1'] = 'Hello, OpenPyXL!'
# 保存工作簿到文件
workbook.save('C:\\Users\\abc\\Documents\\GitHub\\vscodepractice\\python\\openpyxlpractice\\example.xlsx')
workbook.close()  # 关闭工作簿
workbook = openpyxl.load_workbook('C:\\Users\\abc\\Documents\\GitHub\\vscodepractice\\python\\openpyxlpractice\\example.xlsx')
sheet = workbook['MySheet']  # 选择"MySheet"工作表'
print(sheet['A1'].value)  # 输出: Hello, OpenPyXL!
lst=[]
for row in sheet.rows:
    sublst=[]
    for cell in row:
        sublst.append(cell.value)
    lst.append(sublst)
print(lst)